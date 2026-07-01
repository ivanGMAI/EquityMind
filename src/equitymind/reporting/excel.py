"""Excel (.xlsx) workbook export.

Turns an :class:`~equitymind.pipeline.AnalysisReport` into a multi-sheet workbook
that a desk analyst can actually work in — not a static dump:

* **Ranking** — a native Excel table (auto-filter/sortable) with a *live*
  reward-to-risk formula and conditional formatting (colour scales, data bars),
* **Metrics** — the full per-asset numbers (returns, risk-adjusted ratios,
  VaR/CVaR, beta/alpha) in one grid,
* **Portfolio** — the correlation matrix (colour-scaled) and the reference
  allocations,
* **By asset class** — a *pivot-by-formula* summary (``COUNTIF`` / ``AVERAGEIF``)
  that recomputes if the ranking data is edited.

The formulas and conditional formatting are the point: the sheets stay
interactive in Excel, demonstrating the spreadsheet mechanics (formulas, tables,
grouped summaries) rather than pasting frozen values.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..ai.prompts import DISCLAIMER
from ..logging_config import get_logger

if TYPE_CHECKING:  # avoid an import cycle; only needed for type hints
    from ..pipeline import AnalysisReport

logger = get_logger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
_PCT = "0.00%"  # values stored as fractions -> percentage display
_RATIO = "0.00"


class ExcelGenerator:
    """Render :class:`AnalysisReport` objects to .xlsx workbooks."""

    def __init__(self, output_dir: str | Path = "reports") -> None:
        self.output_dir = Path(output_dir)

    # ------------------------------------------------------------------ public
    def write_workbook(self, report: AnalysisReport, *, filename: str | None = None) -> Path:
        """Build and persist the workbook; returns the .xlsx path."""
        self.output_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        wb = self.build_workbook(report)
        out_path = self.output_dir / (filename or f"market_intelligence_{stamp}.xlsx")
        wb.save(out_path)
        logger.info("Excel workbook written to %s", out_path)
        return out_path

    def build_workbook(self, report: AnalysisReport) -> Workbook:
        """Assemble the workbook in memory (no file I/O)."""
        wb = Workbook()
        wb.remove(wb.active)  # drop the default empty sheet
        self._overview_sheet(wb, report)
        n_rows = self._ranking_sheet(wb, report)
        self._metrics_sheet(wb, report)
        if report.portfolio is not None:
            self._portfolio_sheet(wb, report)
        if n_rows:
            self._pivot_sheet(wb, report, n_rows)
        return wb

    # ------------------------------------------------------------------ sheets
    @staticmethod
    def _overview_sheet(wb: Workbook, report: AnalysisReport) -> None:
        ws = wb.create_sheet("Overview")
        ws["A1"] = "EquityMind — Market Intelligence"
        ws["A1"].font = _TITLE_FONT
        ws["A3"] = "Generated"
        ws["B3"] = report.generated_at
        ws["A4"] = "AI provider"
        ws["B4"] = f"{report.ai_provider} ({report.ai_model})"
        ws["A5"] = "Instruments"
        ws["B5"] = len(report.assets)
        ws["A7"] = "Disclaimer"
        ws["A7"].font = Font(bold=True)
        ws["A8"] = DISCLAIMER
        ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A8:H11")
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 40

    def _ranking_sheet(self, wb: Workbook, report: AnalysisReport) -> int:
        """Ranking table with a live reward/risk formula + conditional formats.

        Returns the number of data rows written (0 if no comparison).
        """
        comp = report.comparison
        if comp is None or not comp.entries:
            return 0

        ws = wb.create_sheet("Ranking")
        ws["A1"] = f"Cross-asset ranking (basis: {comp.return_basis})"
        ws["A1"].font = _TITLE_FONT

        headers = [
            "Rank",
            "Ticker",
            "Name",
            "Class",
            "Return",
            "Ann. Vol",
            "Risk (0-100)",
            "Sharpe",
            "Reward/Risk",
            "Trend",
        ]
        header_row = 3
        self._write_header(ws, headers, header_row)

        classes = {a.metrics.ticker: a.metrics.asset_class for a in report.assets.values()}
        sharpes = {
            a.metrics.ticker: (a.metrics.performance.sharpe if a.metrics.performance else None)
            for a in report.assets.values()
        }
        first = header_row + 1
        for i, e in enumerate(comp.entries):
            r = first + i
            ws.cell(r, 1, e.rank)
            ws.cell(r, 2, e.ticker)
            ws.cell(r, 3, e.name)
            ws.cell(r, 4, classes.get(e.ticker, ""))
            ws.cell(
                r, 5, None if e.return_pct is None else e.return_pct / 100.0
            ).number_format = _PCT
            ws.cell(r, 6, e.annualized_volatility_pct / 100.0).number_format = _PCT
            ws.cell(r, 7, e.risk_score)
            ws.cell(r, 8, sharpes.get(e.ticker)).number_format = _RATIO
            # Live formula: reward-to-risk = return / annualised volatility.
            ws.cell(r, 9, f'=IFERROR(E{r}/F{r},"")').number_format = _RATIO
            ws.cell(r, 10, e.trend)
        last = first + len(comp.entries) - 1

        # Summary formulas beneath the table.
        summ = last + 2
        ws.cell(summ, 4, "Average").font = Font(bold=True)
        ws.cell(summ, 5, f"=AVERAGE(E{first}:E{last})").number_format = _PCT
        ws.cell(summ, 6, f"=AVERAGE(F{first}:F{last})").number_format = _PCT
        ws.cell(summ, 7, f"=AVERAGE(G{first}:G{last})").number_format = _RATIO
        ws.cell(summ, 9, f"=AVERAGE(I{first}:I{last})").number_format = _RATIO

        self._make_table(ws, "RankingTable", header_row, last, len(headers))
        # Conditional formatting: risk colour scale, return data bars.
        ws.conditional_formatting.add(
            f"G{first}:G{last}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="63BE7B",
                mid_type="num",
                mid_value=50,
                mid_color="FFEB84",
                end_type="num",
                end_value=100,
                end_color="F8696B",
            ),
        )
        ws.conditional_formatting.add(
            f"E{first}:E{last}", DataBarRule(start_type="min", end_type="max", color="5B9BD5")
        )
        self._autosize(ws, {1: 6, 2: 10, 3: 26, 4: 10, 10: 12})
        return len(comp.entries)

    def _metrics_sheet(self, wb: Workbook, report: AnalysisReport) -> None:
        ws = wb.create_sheet("Metrics")
        ws["A1"] = "Per-instrument metrics"
        ws["A1"].font = _TITLE_FONT
        headers = [
            "Ticker",
            "Class",
            "Last",
            "1d",
            "7d",
            "30d",
            "Cumulative",
            "CAGR",
            "Ann. Vol",
            "Sharpe",
            "Sortino",
            "Calmar",
            "Max DD",
            "VaR (hist)",
            "CVaR (hist)",
            "Beta",
            "Alpha",
            "Corr",
        ]
        header_row = 3
        self._write_header(ws, headers, header_row)

        r = header_row
        for analysis in report.assets.values():
            r += 1
            p = analysis.metrics.to_payload()
            rets, vol, risk = p["returns_pct"], p["volatility"], p["risk"]
            perf, tail = p.get("performance", {}), p.get("tail_risk", {})
            bench = p.get("benchmark") or {}
            row = [
                p["ticker"],
                p["asset_class"],
                p["last_price"],
                _frac(rets.get("1d")),
                _frac(rets.get("7d")),
                _frac(rets.get("30d")),
                _frac(p["cumulative_return_pct"]),
                _frac(perf.get("annualized_return_pct")),
                _frac(vol.get("annualized_pct")),
                perf.get("sharpe"),
                perf.get("sortino"),
                perf.get("calmar"),
                _frac(risk.get("max_drawdown_pct")),
                _frac(tail.get("historical_var_pct")),
                _frac(tail.get("historical_cvar_pct")),
                bench.get("beta"),
                _frac(bench.get("alpha_annual_pct")),
                bench.get("correlation"),
            ]
            for c, value in enumerate(row, start=1):
                cell = ws.cell(r, c, value)
                if c in (4, 5, 6, 7, 8, 9, 13, 14, 15, 17):
                    cell.number_format = _PCT
                elif c in (10, 11, 12, 16, 18):
                    cell.number_format = _RATIO
        self._make_table(ws, "MetricsTable", header_row, r, len(headers))
        self._autosize(ws, {1: 10, 2: 10})

    def _portfolio_sheet(self, wb: Workbook, report: AnalysisReport) -> None:
        p = report.portfolio
        assert p is not None
        ws = wb.create_sheet("Portfolio")
        ws["A1"] = "Portfolio analytics"
        ws["A1"].font = _TITLE_FONT
        ws["A2"] = (
            f"{len(p.tickers)} instruments · {p.observations} obs · "
            f"avg correlation {p.average_correlation:+.2f} · rf {p.risk_free_rate * 100:.1f}%"
        )

        # Correlation matrix with a diverging colour scale.
        start = 4
        ws.cell(start, 1, "Correlation").font = Font(bold=True)
        for j, t in enumerate(p.tickers, start=2):
            ws.cell(start, j, t).font = _HEADER_FONT
            ws.cell(start, j).fill = _HEADER_FILL
        for i, row_t in enumerate(p.tickers, start=1):
            rr = start + i
            ws.cell(rr, 1, row_t).font = Font(bold=True)
            for j, col_t in enumerate(p.tickers, start=2):
                ws.cell(rr, j, p.correlation[row_t][col_t]).number_format = _RATIO
        last_col = get_column_letter(1 + len(p.tickers))
        ws.conditional_formatting.add(
            f"B{start + 1}:{last_col}{start + len(p.tickers)}",
            ColorScaleRule(
                start_type="num",
                start_value=-1,
                start_color="F8696B",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="num",
                end_value=1,
                end_color="63BE7B",
            ),
        )

        # Reference allocations table.
        arow = start + len(p.tickers) + 3
        ws.cell(arow, 1, "Reference allocations").font = Font(bold=True)
        header_row = arow + 1
        headers = ["Allocation", "Exp. return", "Volatility", "Sharpe", *p.tickers]
        self._write_header(ws, headers, header_row)
        r = header_row
        for alloc in p.allocations.values():
            r += 1
            ws.cell(r, 1, alloc.label)
            ws.cell(r, 2, _frac(alloc.expected_return_pct)).number_format = _PCT
            ws.cell(r, 3, _frac(alloc.volatility_pct)).number_format = _PCT
            ws.cell(r, 4, alloc.sharpe).number_format = _RATIO
            for j, t in enumerate(p.tickers, start=5):
                ws.cell(r, j, _frac(alloc.weights.get(t))).number_format = _PCT
        self._autosize(ws, {1: 26})

    def _pivot_sheet(self, wb: Workbook, report: AnalysisReport, n_rows: int) -> None:
        """A 'pivot table' built from AVERAGEIF/COUNTIF formulas over Ranking."""
        ws = wb.create_sheet("By Asset Class")
        ws["A1"] = "Summary by asset class (formula-driven pivot)"
        ws["A1"].font = _TITLE_FONT
        headers = ["Asset class", "Count", "Avg return", "Avg risk", "Avg reward/risk"]
        header_row = 3
        self._write_header(ws, headers, header_row)

        classes: list[str] = []
        for a in report.assets.values():
            cls = a.metrics.asset_class
            if cls not in classes:
                classes.append(cls)

        first = 4  # ranking data starts at row 4
        last = first + n_rows - 1
        rng_class = f"Ranking!$D${first}:$D${last}"
        rng_ret = f"Ranking!$E${first}:$E${last}"
        rng_risk = f"Ranking!$G${first}:$G${last}"
        rng_rr = f"Ranking!$I${first}:$I${last}"
        r = header_row
        for cls in classes:
            r += 1
            ws.cell(r, 1, cls)
            ws.cell(r, 2, f"=COUNTIF({rng_class},A{r})")
            ws.cell(
                r, 3, f'=IFERROR(AVERAGEIF({rng_class},A{r},{rng_ret}),"")'
            ).number_format = _PCT
            ws.cell(
                r, 4, f'=IFERROR(AVERAGEIF({rng_class},A{r},{rng_risk}),"")'
            ).number_format = _RATIO
            ws.cell(
                r, 5, f'=IFERROR(AVERAGEIF({rng_class},A{r},{rng_rr}),"")'
            ).number_format = _RATIO
        # Grand total row.
        r += 1
        ws.cell(r, 1, "All").font = Font(bold=True)
        ws.cell(r, 2, f"=SUM(B{header_row + 1}:B{r - 1})")
        ws.cell(r, 3, f"=AVERAGE({rng_ret})").number_format = _PCT
        ws.cell(r, 4, f"=AVERAGE({rng_risk})").number_format = _RATIO

        # A small bar chart of average return by class.
        chart = BarChart()
        chart.title = "Average return by asset class"
        chart.type = "col"
        data = Reference(ws, min_col=3, min_row=header_row, max_row=r - 1)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=r - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "G3")
        self._autosize(ws, {1: 16})

    # ------------------------------------------------------------------ helpers
    @staticmethod
    def _write_header(ws, headers: list[str], row: int) -> None:
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row, c, h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _make_table(ws, name: str, header_row: int, last_row: int, n_cols: int) -> None:
        if last_row <= header_row:
            return
        ref = f"A{header_row}:{get_column_letter(n_cols)}{last_row}"
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

    @staticmethod
    def _autosize(ws, widths: dict) -> None:
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width


def _frac(pct: float | None) -> float | None:
    """Convert a percentage value back to a fraction for %-formatted cells."""
    return None if pct is None else pct / 100.0
